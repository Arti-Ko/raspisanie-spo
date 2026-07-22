from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(bold=True, size=11, color="1F3864")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
OVER_LIMIT_FILL = PatternFill("solid", fgColor="F8CBAD")
THIN = Side(style="thin", color="B7B7B7")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(worksheet, row: int, first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER


def style_data_row(
    worksheet, row: int, first_col: int, last_col: int, shaded: bool = False
) -> None:
    for col in range(first_col, last_col + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.border = CELL_BORDER
        cell.alignment = LEFT
        if shaded:
            cell.fill = SUBHEADER_FILL
