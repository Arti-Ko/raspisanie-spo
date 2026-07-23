from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from app.export.excel_styles import (
    SUBTITLE_FONT,
    TITLE_FONT,
    style_data_row,
    style_header_row,
)
from app.export.schedule_layout import DAY_LABELS, blocks_for_week, build_block_rows
from app.repositories.schedule import ScheduleEntry
from app.repositories.text_format import abbreviate_name

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def export_schedule(
    group_name: str,
    week_label: str,
    date_range_label: str,
    entries: list[ScheduleEntry],
    includes_saturday: bool,
    file_path: str,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Расписание"
    entry_by_slot = {(e.day_of_week, e.pair_number): e for e in entries}
    blocks = blocks_for_week(includes_saturday)
    max_columns = 2 + 2 * max(len(days) for days in blocks)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_columns)
    sheet.cell(
        row=1, column=1, value=f"Расписание группы {group_name} — {week_label}"
    ).font = TITLE_FONT

    row = 2
    if date_range_label:
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=max_columns
        )
        sheet.cell(row=row, column=1, value=f"Даты: {date_range_label}")
        row += 1

    row += 1
    for days in blocks:
        row = write_block(sheet, row, days, entry_by_slot)
        sheet.row_breaks.append(Break(id=row))
        row += 2

    sheet.column_dimensions["A"].width = 6
    sheet.column_dimensions["B"].width = 13
    for col_index in range(3, max_columns + 1, 2):
        sheet.column_dimensions[get_column_letter(col_index)].width = 26
        sheet.column_dimensions[get_column_letter(col_index + 1)].width = 4

    workbook.save(file_path)


def write_block(
    sheet,
    start_row: int,
    days: tuple[int, ...],
    entry_by_slot: dict,
) -> int:
    block_columns = 2 + 2 * len(days)
    title = " – ".join(DAY_LABELS[day] for day in (days[0], days[-1]))
    sheet.merge_cells(
        start_row=start_row, start_column=1, end_row=start_row, end_column=block_columns
    )
    sheet.cell(row=start_row, column=1, value=title).font = SUBTITLE_FONT

    header_row = start_row + 1
    sheet.cell(row=header_row, column=1, value="Пара")
    sheet.cell(row=header_row, column=2, value="Время")
    for col_offset, day in enumerate(days):
        col = 3 + col_offset * 2
        sheet.cell(row=header_row, column=col, value=DAY_LABELS[day])
        sheet.cell(row=header_row, column=col + 1, value="Каб")
    style_header_row(sheet, header_row, 1, block_columns)

    rows = build_block_rows(days)
    pair_merge_start: dict[int, int] = {}
    for row_offset, spec in enumerate(rows):
        row = header_row + 1 + row_offset
        sheet.cell(row=row, column=2, value=spec.time_label)

        if spec.is_zero_period:
            sheet.cell(row=row, column=1, value="0")
        elif spec.starts_pair:
            pair_merge_start[spec.pair_number] = row
            sheet.cell(row=row, column=1, value=str(spec.pair_number))
        else:
            first_row = pair_merge_start[spec.pair_number]
            sheet.merge_cells(
                start_row=first_row, start_column=1, end_row=row, end_column=1
            )

        for col_offset, day in enumerate(days):
            content_col = 3 + col_offset * 2
            room_col = content_col + 1
            entry = entry_by_slot.get((day, spec.pair_number))
            if spec.is_zero_period:
                _write_entry(sheet, row, row, content_col, room_col, entry)
            elif spec.starts_pair:
                _write_entry(sheet, row, row + 1, content_col, room_col, entry)
        style_data_row(sheet, row, 1, block_columns)
        sheet.row_dimensions[row].height = 30

    return header_row + len(rows)


def _write_entry(sheet, row_start, row_end, content_col, room_col, entry) -> None:
    if row_end > row_start:
        sheet.merge_cells(
            start_row=row_start,
            start_column=content_col,
            end_row=row_end,
            end_column=content_col,
        )
        sheet.merge_cells(
            start_row=row_start,
            start_column=room_col,
            end_row=row_end,
            end_column=room_col,
        )
    if entry is None:
        return
    subject_cell = sheet.cell(
        row=row_start, column=content_col, value=_entry_text(entry)
    )
    subject_cell.alignment = CENTER
    room_cell = sheet.cell(row=row_start, column=room_col, value=entry.room or "")
    room_cell.alignment = CENTER


def _entry_text(entry: ScheduleEntry) -> str:
    name = abbreviate_name(entry.effective_teacher_name)
    text = f"{entry.subject_name}\n{name}"
    if entry.substitute_teacher_id:
        text += " (замена)"
    return text
