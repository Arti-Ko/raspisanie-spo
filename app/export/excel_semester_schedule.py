from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from app.export.excel_schedule import write_block
from app.export.excel_styles import SUBTITLE_FONT, TITLE_FONT
from app.export.schedule_layout import blocks_for_week
from app.repositories.calendar_weeks import CalendarWeek
from app.repositories.groups import Group
from app.repositories.schedule import ScheduleEntry
from app.repositories.teachers import list_teachers

MAX_SHEET_NAME = 31


def export_semester_schedule(
    groups: list[Group],
    weeks: list[CalendarWeek],
    semester_label: str,
    entries_by_group_week: dict[tuple[int, int], list[ScheduleEntry]],
    file_path: str,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    teacher_colors = {t.id: t.color for t in list_teachers()}
    max_columns = 2 + 2 * 3

    for group in groups:
        sheet = workbook.create_sheet(title=group.name[:MAX_SHEET_NAME])
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=max_columns
        )
        sheet.cell(
            row=1, column=1, value=f"Расписание группы {group.name} — {semester_label}"
        ).font = TITLE_FONT

        row = 3
        for week in weeks:
            entries = entries_by_group_week.get((group.id, week.id), [])
            entry_by_slot = {(e.day_of_week, e.pair_number): e for e in entries}
            week_title = f"Неделя {week.week_number}"
            if week.date_range_label:
                week_title += f" ({week.date_range_label})"
            sheet.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=max_columns
            )
            sheet.cell(row=row, column=1, value=week_title).font = SUBTITLE_FONT
            row += 1

            for days in blocks_for_week(week.includes_saturday):
                row = write_block(sheet, row, days, entry_by_slot, teacher_colors)
                sheet.row_breaks.append(Break(id=row))
                row += 2
            row += 1

        sheet.column_dimensions["A"].width = 6
        sheet.column_dimensions["B"].width = 13
        for col_index in range(3, max_columns + 1, 2):
            sheet.column_dimensions[get_column_letter(col_index)].width = 24
            sheet.column_dimensions[get_column_letter(col_index + 1)].width = 6

    workbook.save(file_path)
