from openpyxl import Workbook

from app.export.excel_styles import (
    SUBTITLE_FONT,
    TITLE_FONT,
    style_data_row,
    style_header_row,
)
from app.repositories.reports import TeacherReport

WEEKLY_HEADERS = ["Полугодие", "Неделя", "Группа", "Предмет", "Часов", "Замена"]
SUB_HEADERS = ["Полугодие", "Неделя", "Группа", "Предмет", "Часов", "Кто"]


def export_teacher_report(report: TeacherReport, file_path: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Справка о часах"

    sheet.merge_cells("A1:F1")
    sheet["A1"] = "СПРАВКА О ПРОЧИТАННЫХ ЧАСАХ"
    sheet["A1"].font = TITLE_FONT

    sheet["A2"] = f"Преподаватель: {report.teacher_name}"
    sheet["A2"].font = SUBTITLE_FONT

    sheet["A4"] = "По тарификации"
    sheet["B4"] = f"I полугодие: {report.plan_hours[1]} ч."
    sheet["C4"] = f"II полугодие: {report.plan_hours[2]} ч."
    sheet["D4"] = f"Всего: {report.plan_total} ч."

    sheet["A5"] = "Прочитано фактически"
    sheet["B5"] = f"I полугодие: {report.actual_hours[1]} ч."
    sheet["C5"] = f"II полугодие: {report.actual_hours[2]} ч."
    sheet["D5"] = f"Всего: {report.actual_total} ч."

    sheet["A6"] = "Разница (план − факт)"
    sheet["B6"] = f"{report.difference} ч."

    row = 8
    sheet.cell(
        row=row, column=1, value="Понедельная разбивка прочитанных часов"
    ).font = SUBTITLE_FONT
    row += 1
    for col, title in enumerate(WEEKLY_HEADERS, start=1):
        sheet.cell(row=row, column=col, value=title)
    style_header_row(sheet, row, 1, len(WEEKLY_HEADERS))
    row += 1
    for line in report.weekly_lines:
        sheet.cell(row=row, column=1, value=f"{line.semester}-е")
        sheet.cell(row=row, column=2, value=line.week_number)
        sheet.cell(row=row, column=3, value=line.group_name)
        sheet.cell(row=row, column=4, value=line.subject_name)
        sheet.cell(row=row, column=5, value=line.hours)
        sheet.cell(row=row, column=6, value="да" if line.is_substitute else "")
        style_data_row(sheet, row, 1, len(WEEKLY_HEADERS))
        row += 1

    row += 1
    row = _write_substitution_table(
        sheet, row, "Кто заменял этого преподавателя", report.covered_by_others
    )
    row += 1
    _write_substitution_table(
        sheet, row, "Кого заменял этот преподаватель", report.covered_for_others
    )

    sheet["A" + str(row + len(report.covered_for_others) + 4)] = (
        "Подпись преподавателя ____________________"
    )
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 26
    sheet.column_dimensions["E"].width = 10
    sheet.column_dimensions["F"].width = 22

    workbook.save(file_path)


def _write_substitution_table(sheet, row: int, title: str, lines) -> int:
    sheet.cell(row=row, column=1, value=title).font = SUBTITLE_FONT
    row += 1
    for col, header in enumerate(SUB_HEADERS, start=1):
        sheet.cell(row=row, column=col, value=header)
    style_header_row(sheet, row, 1, len(SUB_HEADERS))
    row += 1
    if not lines:
        sheet.cell(row=row, column=1, value="— нет —")
        style_data_row(sheet, row, 1, len(SUB_HEADERS))
        row += 1
        return row
    for line in lines:
        sheet.cell(row=row, column=1, value=f"{line.semester}-е")
        sheet.cell(row=row, column=2, value=line.week_number)
        sheet.cell(row=row, column=3, value=line.group_name)
        sheet.cell(row=row, column=4, value=line.subject_name)
        sheet.cell(row=row, column=5, value=line.hours)
        sheet.cell(row=row, column=6, value=line.other_teacher_name)
        style_data_row(sheet, row, 1, len(SUB_HEADERS))
        row += 1
    return row
